''' openpyxl - ler e alterar dados de uma planilha '''

from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.cell import Cell
from openpyxl.worksheet.worksheet import Worksheet

ROOT_FOLDER = Path(__file__).parent
WORKBOOK_PATH = ROOT_FOLDER / 'workbook.xlsx'

# Carregando um arquivo do excel
workbook: Workbook = load_workbook(WORKBOOK_PATH)

# Nome para a planilha
sheet_name = 'Minha planilha'

# Selecionou a planilha
worksheet: Worksheet = workbook[sheet_name]

row: tuple[Cell]
for row in worksheet.iter_rows(min_row=2):  # type: ignore
    for cell in row:
        print(cell.value, end='\t')

        ''' alterando idade de Maria no loop for'''
        # if cell.value == 'Maria':
        #     worksheet.cell(cell.row, 2, 23)

    print()

''' alterando idade de Maria'''
# worksheet['B3'].value = 14

workbook.save(WORKBOOK_PATH)
